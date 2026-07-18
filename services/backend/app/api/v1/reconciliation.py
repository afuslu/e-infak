import csv
import hashlib
import io
import unicodedata
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import Permission, require_permission
from app.api.v1.checkout import _finalize_paid_order, _load_order
from app.core.db import get_db
from app.models.automation import BankStatementImport, BankTransaction, ReconciliationStatus
from app.models.payment import PaymentAttempt, PaymentOrder, PaymentOrderStatus
from app.models.user import User

router = APIRouter(tags=["reconciliation"])


def _amount_to_cents(value) -> int:
    text = str(value or "").strip().replace("₺", "").replace("TL", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return int((Decimal(text) * 100).quantize(Decimal("1")))
    except InvalidOperation as exc:
        raise ValueError(f"Geçersiz tutar: {value}") from exc


def _parse_date(value) -> datetime:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(str(value or "").strip(), fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    raise ValueError(f"Geçersiz tarih: {value}")


def _normalize_row(row: dict, index: int) -> dict:
    def normalize_key(value) -> str:
        folded = unicodedata.normalize("NFKD", str(value).strip().casefold())
        return "".join(char for char in folded if not unicodedata.combining(char)).replace("ı", "i")
    normalized = {normalize_key(k): v for k, v in row.items() if k is not None}
    def pick(*keys):
        return next((normalized[k] for k in keys if normalized.get(k) not in (None, "")), None)
    external_id = pick("işlem no", "islem no", "transaction_id", "referans", "reference")
    if not external_id:
        row_fingerprint = "|".join(f"{key}={row[key]}" for key in sorted(row, key=str))
        external_id = f"ROW-{hashlib.sha256(row_fingerprint.encode()).hexdigest()[:24]}"
    return {
        "external_id": str(external_id).strip(),
        "booked_at": _parse_date(pick("tarih", "date", "işlem tarihi", "islem tarihi")),
        "amount_cents": _amount_to_cents(pick("tutar", "amount", "alacak")),
        "currency": str(pick("para birimi", "currency", "döviz") or "TRY").upper(),
        "sender_name": str(pick("gönderen", "gonderen", "sender", "hesap sahibi") or "").strip() or None,
        "description": str(pick("açıklama", "aciklama", "description") or "").strip(),
        "raw_data": {str(k): str(v) for k, v in row.items() if k is not None and v is not None},
    }


def _read_rows(content: bytes, filename: str) -> list[dict]:
    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        return [_normalize_row(dict(zip(headers, row)), index) for index, row in enumerate(values, 2)]
    text = content.decode("utf-8-sig")
    dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [_normalize_row(row, index) for index, row in enumerate(reader, 2)]


async def _auto_match(db: AsyncSession, transaction: BankTransaction, user_id: UUID):
    result = await db.execute(
        select(PaymentOrder).where(
            PaymentOrder.organization_id == transaction.organization_id,
            PaymentOrder.payment_method == "bank_transfer",
            PaymentOrder.status == PaymentOrderStatus.AWAITING_TRANSFER,
            PaymentOrder.total_cents == transaction.amount_cents,
            PaymentOrder.currency == transaction.currency,
        )
    )
    description = (transaction.description or "").upper()
    candidates = [o for o in result.scalars().all() if o.transfer_reference and o.transfer_reference.upper() in description]
    if len(candidates) > 1:
        transaction.status = ReconciliationStatus.AMBIGUOUS
        return
    if not candidates:
        return
    order = await _load_order(db, candidates[0].id)
    attempt = PaymentAttempt(
        order_id=order.id, provider="bank_transfer", pg_tran_id=transaction.external_id,
        response_code="00", response_message="Banka ekstresiyle otomatik eşleşti",
        safe_response={"bank_transaction_id": str(transaction.id)},
    )
    db.add(attempt)
    transaction.status = ReconciliationStatus.MATCHED
    transaction.payment_order_id = order.id
    transaction.matched_by_id = user_id
    transaction.matched_at = datetime.now(timezone.utc)
    await db.flush()
    await _finalize_paid_order(db, order, attempt, commit=False)


@router.post("/admin/bank-statements/import")
async def import_bank_statement(
    request: Request,
    file: UploadFile = File(...),
    preview: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.FINANCE_MANAGE)),
):
    if not file.filename or not file.filename.lower().endswith((".csv", ".xlsx")):
        raise HTTPException(status_code=400, detail="Yalnız CSV veya XLSX kabul edilir")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Dosya en fazla 10 MB olabilir")
    try:
        rows = _read_rows(content, file.filename)
    except (ValueError, StopIteration, UnicodeDecodeError, csv.Error) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if preview:
        return {"preview": True, "row_count": len(rows), "rows": rows[:20]}
    org_id = UUID(request.state.organization_id)
    file_hash = hashlib.sha256(content).hexdigest()
    duplicate = await db.execute(select(BankStatementImport).where(
        BankStatementImport.organization_id == org_id, BankStatementImport.file_hash == file_hash
    ))
    if duplicate.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Bu ekstre daha önce içe aktarıldı")
    statement = BankStatementImport(
        organization_id=org_id, file_hash=file_hash, filename=file.filename,
        row_count=len(rows), imported_by_id=current_user.id,
    )
    db.add(statement)
    await db.flush()
    imported_count = 0
    skipped_count = 0
    for row in rows:
        existing = await db.execute(select(BankTransaction.id).where(
            BankTransaction.organization_id == org_id,
            BankTransaction.external_id == row["external_id"],
        ))
        if existing.scalar_one_or_none():
            skipped_count += 1
            continue
        transaction = BankTransaction(organization_id=org_id, import_id=statement.id, **row)
        db.add(transaction)
        await db.flush()
        await _auto_match(db, transaction, current_user.id)
        imported_count += 1
    await db.commit()
    return {
        "import_id": statement.id,
        "row_count": len(rows),
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "status": "imported",
    }


@router.get("/admin/reconciliation")
async def list_reconciliation(
    request: Request,
    status: ReconciliationStatus | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.FINANCE_MANAGE)),
):
    query = select(BankTransaction).where(BankTransaction.organization_id == UUID(request.state.organization_id))
    if status:
        query = query.where(BankTransaction.status == status)
    result = await db.execute(query.order_by(BankTransaction.booked_at.desc()).limit(500))
    return result.scalars().all()


@router.post("/admin/reconciliation/{transaction_id}/resolve")
async def resolve_reconciliation(
    request: Request,
    transaction_id: UUID,
    payment_order_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.FINANCE_MANAGE)),
):
    org_id = UUID(request.state.organization_id)
    result = await db.execute(select(BankTransaction).where(
        BankTransaction.id == transaction_id, BankTransaction.organization_id == org_id
    ))
    transaction = result.scalar_one_or_none()
    order = await _load_order(db, payment_order_id)
    if not transaction or not order or order.organization_id != org_id:
        raise HTTPException(status_code=404, detail="Hareket veya ödeme bulunamadı")
    if order.payment_method != "bank_transfer" or order.status != PaymentOrderStatus.AWAITING_TRANSFER:
        raise HTTPException(status_code=409, detail="Ödeme eşleştirmeye uygun değil")
    if transaction.amount_cents != order.total_cents or transaction.currency != order.currency:
        raise HTTPException(status_code=409, detail="Tutar veya para birimi eşleşmiyor")
    attempt = PaymentAttempt(
        order_id=order.id, provider="bank_transfer", pg_tran_id=transaction.external_id,
        response_code="00", response_message="Muhasebe tarafından eşleştirildi",
        safe_response={"bank_transaction_id": str(transaction.id)},
    )
    db.add(attempt)
    transaction.status = ReconciliationStatus.MATCHED
    transaction.payment_order_id = order.id
    transaction.matched_by_id = current_user.id
    transaction.matched_at = datetime.now(timezone.utc)
    await db.flush()
    donation_ids = await _finalize_paid_order(db, order, attempt)
    return {"status": "matched", "donation_count": len(donation_ids)}
